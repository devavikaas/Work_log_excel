
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import datetime

# === Configuration ===
start_date = datetime(2025, 6, 1)
end_date = datetime(2025, 6, 30)
date_range = pd.date_range(start=start_date, end=end_date)

columns = [
    "Start Time", "End Time", "Duration (minutes)",
    "Type", "Task Name", "Description / Subject", "Remarks"
]

header_colors = [
    "FBE5D6", "D9EAD3", "D9D2E9", "CFE2F3", "FFF2CC", "F4CCCC", "EAD1DC"
]
header_fills = [PatternFill(start_color=c, end_color=c, fill_type="solid") for c in header_colors]
header_font = Font(bold=True, size=16)
cell_font = Font(size=14)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
summary_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
tab_colors = [
    "1072BA", "A9D08E", "FFD966", "D9D2E9", "F4CCCC", "B4C7E7", "FFE699",
    "C9DAF8", "D5A6BD", "B6D7A8", "EAD1DC", "F9CB9C", "B7E1CD", "F6B26B",
    "D0E0E3", "FCE5CD", "A2C4C9", "C27BA0", "A4C2F4", "D5A6BD", "B6D7A8",
    "D9D2E9", "C9DAF8", "F4CCCC", "A2C4C9", "F9CB9C", "FFE699", "D5A6BD",
    "C27BA0", "B4C7E7"
]

file_path = "/content/UPSC_Study_Log_June_2025.xlsx"

# === Step 1: Create base workbook with day number sheets ===
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    for i, date in enumerate(date_range, 1):
        sheet_name = str(i)
        df = pd.DataFrame([[""] * len(columns)], columns=columns)
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# === Step 2: Load workbook and style each sheet ===
wb = openpyxl.load_workbook(file_path)

for i, sheet_name in enumerate(wb.sheetnames):
    if sheet_name == "Summary" or sheet_name == "Home":
        continue

    ws = wb[sheet_name]
    actual_date = date_range[i].strftime("%B %d, %Y")

    # Add date title in A1
    ws.insert_rows(1)
    ws["A1"] = actual_date
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Header styling in row 2
    for col_idx, (fill, header) in enumerate(zip(header_fills, columns), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format cells
    for row in ws.iter_rows(min_row=3, max_row=102, min_col=1, max_col=len(columns)):
        for cell in row:
            cell.font = cell_font
            cell.alignment = wrap_alignment

    # Time + Duration formulas
    for row in range(3, 103):
       # Format Start Time and End Time cells
       ws[f"A{row}"].number_format = "h:mm:ss AM/PM"
       ws[f"B{row}"].number_format = "h:mm:ss AM/PM"

    # Autochain: A4 gets "=B3", A5 gets "=B4", ...
       if row > 3:
        ws[f"A{row}"] = f"=B{row - 1}"

    # Duration calculation
       ws[f"C{row}"] = (
        f"=IF(AND(ISNUMBER(B{row}), ISNUMBER(A{row})), ROUND((B{row}-A{row})*1440, 0), \"\")"
       )
       ws[f"C{row}"].number_format = "0"


    # Dropdown in Type column
    dv = DataValidation(type="list", formula1='"Task,Break"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("D3:D102")

    # Add table
    end_col_letter = get_column_letter(len(columns))
    table_range = f"A2:{end_col_letter}102"
    table_name = "T_" + sheet_name
    table = Table(displayName=table_name, ref=table_range)
    table_style = TableStyleInfo(name="TableStyleMedium6", showRowStripes=True)
    table.tableStyleInfo = table_style
    ws.add_table(table)

    # Summary block
    ws["I2"] = "Total Study Time"
    ws["J2"] = "=SUMIF(D3:D102, \"Task\", C3:C102)/1440"
    ws["I3"] = "Total Break Time"
    ws["J3"] = "=SUMIF(D3:D102, \"Break\", C3:C102)/1440"
    ws["I4"] = "Focus Ratio"
    ws["J4"] = "=IF((J2+J3)=0, 0, J2/(J2+J3))"

    ws["J2"].number_format = "hh:mm:ss"
    ws["J3"].number_format = "hh:mm:ss"
    ws["J4"].number_format = "0.00%"

    for cell in ["I2", "I3", "I4"]:
        ws[cell].font = header_font
        ws[cell].fill = summary_fill
        ws[cell].alignment = Alignment(horizontal="left", wrap_text=True)

    for cell in ["J2", "J3", "J4"]:
        ws[cell].font = cell_font
        ws[cell].alignment = wrap_alignment

    # Column width auto adjust
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(len(str(ws[f"{col_letter}{row}"].value) or "") for row in range(1, 103))
        ws.column_dimensions[col_letter].width = max_length + 2

    ws.sheet_properties.tabColor = tab_colors[i % len(tab_colors)]

# === Step 3: Create Summary Sheet ===
summary = wb.create_sheet("Summary")
headers = ["Day", "Total Study Time", "Total Break Time", "Focus Ratio"]

for col_idx, header in enumerate(headers, 1):
    cell = summary.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = summary_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for i in range(1, len(date_range) + 1):
    summary[f"A{i+1}"] = f"Day {i}"
    ref = f"'{i}'"
    summary[f"B{i+1}"] = f"=SUMIF({ref}!D3:D102, \"Task\", {ref}!C3:C102)/1440"
    summary[f"C{i+1}"] = f"=SUMIF({ref}!D3:D102, \"Break\", {ref}!C3:C102)/1440"
    summary[f"D{i+1}"] = f"=IF((B{i+1}+C{i+1})=0, 0, B{i+1}/(B{i+1}+C{i+1}))"

    for col in "BCD":
        summary[f"{col}{i+1}"].number_format = {
            "B": "hh:mm:ss", "C": "hh:mm:ss", "D": "0.00%"
        }[col]

    for col_idx in range(1, 5):
        summary.cell(row=i+1, column=col_idx).font = cell_font
        summary.cell(row=i+1, column=col_idx).alignment = wrap_alignment

# Adjust summary columns
for col_idx in range(1, 5):
    col_letter = get_column_letter(col_idx)
    max_len = max(len(str(summary[f"{col_letter}{row}"].value) or "") for row in range(1, len(date_range)+2))
    summary.column_dimensions[col_letter].width = max_len + 2

# Add table to summary
summary_range = f"A1:D{len(date_range)+1}"
summary_table = Table(displayName="SummaryTable", ref=summary_range)
summary_style = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
summary_table.tableStyleInfo = summary_style
summary.add_table(summary_table)

# === Step 4: Create Home Sheet with hyperlinks ===
home = wb.create_sheet("Home", 0)
home["A1"] = "Study Log - June 2025"
home["A1"].font = Font(bold=True, size=20)
home["A1"].alignment = Alignment(horizontal="center")
home.merge_cells("A1:B1")

for i in range(1, len(date_range) + 1):
    cell = home.cell(row=i + 1, column=1, value=f"Day {i}")
    cell.font = Font(color="0000FF", underline="single", size=14)
    cell.hyperlink = f"#{i}!A1"

home.column_dimensions["A"].width = 20

# Save workbook
wb.save(file_path)

# Optional (for Colab):
from google.colab import files
files.download(file_path)
